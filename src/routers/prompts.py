import io
import json
import pandas as pd
from pathlib import Path
from fastapi import APIRouter, UploadFile, File, Form, Depends, HTTPException
from fastapi.responses import JSONResponse, StreamingResponse

from src.database.engine import Database
from src.database.repository import Repository
from src.config import load_config
from openpyxl import Workbook
from openpyxl.styles import Font

router = APIRouter()
config = load_config()
database = Database(config.postgres)

def get_repo():
    session = database.get_sync_session()
    try:
        yield Repository(session)
    finally:
        session.close()

@router.get("/api/prompts")
def list_prompt_sets(repo: Repository = Depends(get_repo)):
    sets = repo.get_prompt_sets()
    result = []
    for s in sets:
        result.append({
            "id": s.id,
            "name": s.name,
            "description": s.description,
            "created_at": s.created_at.isoformat() if s.created_at else None,
            "prompt_count": len(s.prompts)
        })
    return result

@router.post("/api/prompts")
async def upload_prompt_set(
    name: str = Form(...),
    description: str = Form(""),
    file: UploadFile = File(...),
    repo: Repository = Depends(get_repo)
):
    try:
        content = await file.read()
        xls = pd.ExcelFile(io.BytesIO(content))
        
        prompts = []
        for sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name)
            # Find columns
            cols = [c.lower() for c in df.columns]
            prompt_col = None
            for c in df.columns:
                if "prompt" in str(c).lower():
                    prompt_col = c
                    break
            
            if not prompt_col:
                continue # Skip sheets without prompt column
                
            for _, row in df.iterrows():
                p_text = str(row[prompt_col]).strip()
                if p_text and p_text.lower() != "nan":
                    prompts.append({
                        "scenario": sheet_name,
                        "prompt_text": p_text
                    })
                    
        if not prompts:
            raise HTTPException(status_code=400, detail="No valid prompts found in Excel")
            
        pset = repo.create_prompt_set(name=name, description=description)
        repo.add_prompts_to_set(pset.id, prompts)
        
        return {"status": "success", "id": pset.id, "prompts_loaded": len(prompts)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.delete("/api/prompts/{pset_id}")
def delete_prompt_set(pset_id: int, repo: Repository = Depends(get_repo)):
    repo.delete_prompt_set(pset_id)
    return {"status": "success"}

@router.get("/api/prompts/template")
def download_template():
    wb = Workbook()
    
    # Scenarios based on our standard default scenarios
    scenarios = ["code_generation", "simple_chat", "rag_retrieval", "reasoning"]
    
    for i, s in enumerate(scenarios):
        if i == 0:
            ws = wb.active
            ws.title = s
        else:
            ws = wb.create_sheet(title=s)
            
        ws["A1"] = "No"
        ws["B1"] = "Prompt"
        ws["A1"].font = Font(bold=True)
        ws["B1"].font = Font(bold=True)
        
        # Add sample data
        ws["A2"] = 1
        ws["B2"] = f"Sample prompt for {s} 1"
        ws["A3"] = 2
        ws["B3"] = f"Sample prompt for {s} 2"
        
        ws.column_dimensions["B"].width = 80
        
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    return StreamingResponse(
        iter([stream.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=prompt_template.xlsx"}
    )
