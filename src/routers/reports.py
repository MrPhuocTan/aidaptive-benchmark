from fastapi import APIRouter, Request, Depends
from fastapi.responses import HTMLResponse, RedirectResponse
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.ext.asyncio import AsyncSession

from src.app import _render, get_db, _database_ready, _db_warning_payload, config
from src.database.repository import AsyncRepository

router = APIRouter()

@router.get("/reports", response_class=HTMLResponse)
async def page_reports(
    request: Request,
    session: AsyncSession = Depends(get_db),
):
    db_ok = _database_ready()
    completed_runs = []

    if db_ok:
        repo = AsyncRepository(session)
        try:
            runs = await repo.list_runs(limit=50)
            completed_runs = [r for r in runs if r.status == "completed"]
        except SQLAlchemyError:
            db_ok = False

    return _render(
        request,
        "reports.html",
        {
            "page": "reports",
            "config": config,
            "runs": completed_runs,
            "db_available": db_ok,
            **(_db_warning_payload() if not db_ok else {}),
        },
    )

@router.get("/reports/{run_id}", response_class=HTMLResponse)
async def page_report_details(
    request: Request,
    run_id: str,
    session: AsyncSession = Depends(get_db),
):
    db_ok = _database_ready()
    report_data = None
    run = None

    if db_ok:
        repo = AsyncRepository(session)
        try:
            run = await repo.get_run(run_id)
            if run:
                report_data = await repo.get_detailed_report_stats(run_id)
                chart_data = await repo.get_dashboard_chart_data(run_id)
                timeline_chart = await repo.get_timeline_chart_data(run_id)
        except SQLAlchemyError:
            db_ok = False

    if not run:
        return RedirectResponse(url="/reports")

    return _render(
        request,
        "report_details.html",
        {
            "page": "reports",
            "config": config,
            "run": run,
            "report_data": report_data,
            "chart_data": chart_data if run else None,
            "timeline_chart": timeline_chart if run else None,
            "db_available": db_ok,
            **(_db_warning_payload() if not db_ok else {}),
        },
    )

@router.get("/reports/{run_id}/download", response_class=HTMLResponse)
async def page_report_download(
    request: Request,
    run_id: str,
    session: AsyncSession = Depends(get_db),
):
    db_ok = _database_ready()
    report_data = None
    run = None

    if db_ok:
        repo = AsyncRepository(session)
        try:
            run = await repo.get_run(run_id)
            if run:
                report_data = await repo.get_detailed_report_stats(run_id)
                chart_data = await repo.get_dashboard_chart_data(run_id)
                timeline_chart = await repo.get_timeline_chart_data(run_id)
        except SQLAlchemyError:
            db_ok = False

    if not run:
        return RedirectResponse(url="/reports")

    response = _render(
        request,
        "report_standalone.html",
        {
            "run": run,
            "report_data": report_data,
            "chart_data": chart_data if run else None,
            "timeline_chart": timeline_chart if run else None,
        },
    )
    response.headers["Content-Disposition"] = f'attachment; filename="aidaptive_report_{run_id}.html"'
    return response


# --------------------------------------------------
# Excel Export — Multi-sheet comprehensive report
# --------------------------------------------------
@router.get("/reports/{run_id}/export/excel")
async def export_excel(
    run_id: str,
    session: AsyncSession = Depends(get_db),
):
    import io
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if not _database_ready():
        return RedirectResponse(url="/reports")

    repo = AsyncRepository(session)
    try:
        run = await repo.get_run(run_id)
        if not run:
            return RedirectResponse(url="/reports")

        results = await repo.get_results_by_run(run_id)
        hardware = await repo.get_hardware_metrics_by_run(run_id)
        comparisons = await repo.get_comparisons_by_run(run_id)
        report_data = await repo.get_detailed_report_stats(run_id)
    except SQLAlchemyError:
        return RedirectResponse(url="/reports")

    wb = Workbook()

    # --- Styles ---
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2B4C80", end_color="2B4C80", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def style_header_row(ws, row_num, col_count):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    # =============================================
    # Sheet 1: Summary
    # =============================================
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 50

    summary_rows = [
        ("Run ID", run.run_id),
        ("Status", run.status),
        ("Test Suite", (run.suite or "").replace("_", " ").title()),
        ("Model", run.model or "N/A"),
        ("Started At", run.started_at.strftime("%Y-%m-%d %H:%M:%S") if run.started_at else "N/A"),
        ("Finished At", run.finished_at.strftime("%Y-%m-%d %H:%M:%S") if run.finished_at else "N/A"),
        ("Duration (s)", round(run.duration_seconds, 1) if run.duration_seconds else "N/A"),
        ("Servers", ", ".join(report_data["metadata"]["servers"]) if report_data else "N/A"),
        ("Tools", ", ".join(report_data["metadata"]["tools"]) if report_data else "N/A"),
        ("Scenarios", ", ".join(report_data["metadata"]["scenarios"]) if report_data else "N/A"),
        ("Total Tests", run.total_tests or 0),
        ("Completed Tests", run.completed_tests or 0),
        ("Failed Tests", run.failed_tests or 0),
        ("Total Prompts Processed", report_data["metadata"]["total_requests"] if report_data else 0),
    ]

    ws_summary.append(["Field", "Value"])
    style_header_row(ws_summary, 1, 2)
    for label, value in summary_rows:
        ws_summary.append([label, value])

    # =============================================
    # Sheet 2: Benchmark Results (all test cases)
    # =============================================
    ws_results = wb.create_sheet("Benchmark Results")

    result_headers = [
        "#", "Timestamp", "Server", "Tool", "Environment", "Scenario", "Model",
        "Concurrency", "TTFT (ms)", "TPOT (ms)", "ITL (ms)", "TPS",
        "RPS", "P50 (ms)", "P95 (ms)", "P99 (ms)",
        "Total Tokens", "Total Requests", "Successful", "Failed",
        "Error Rate (%)", "Goodput",
    ]
    ws_results.append(result_headers)
    style_header_row(ws_results, 1, len(result_headers))

    for idx, r in enumerate(results, 1):
        error_pct = round(r.error_rate * 100, 2) if r.error_rate is not None else None
        ws_results.append([
            idx,
            r.timestamp.strftime("%Y-%m-%d %H:%M:%S") if r.timestamp else "",
            r.server, r.tool, r.environment, r.scenario, r.model,
            r.concurrency,
            round(r.ttft_ms, 3) if r.ttft_ms is not None else None,
            round(r.tpot_ms, 3) if r.tpot_ms is not None else None,
            round(r.itl_ms, 3) if r.itl_ms is not None else None,
            round(r.tps, 3) if r.tps is not None else None,
            round(r.rps, 3) if r.rps is not None else None,
            round(r.latency_p50_ms, 3) if r.latency_p50_ms is not None else None,
            round(r.latency_p95_ms, 3) if r.latency_p95_ms is not None else None,
            round(r.latency_p99_ms, 3) if r.latency_p99_ms is not None else None,
            r.total_tokens,
            r.total_requests,
            r.successful_requests,
            r.failed_requests,
            error_pct,
            round(r.goodput, 3) if r.goodput is not None else None,
        ])

    # Auto-width for results sheet
    for col_idx in range(1, len(result_headers) + 1):
        ws_results.column_dimensions[ws_results.cell(1, col_idx).column_letter].width = max(
            len(str(result_headers[col_idx - 1])) + 2, 12
        )

    # =============================================
    # Sheet 3: Hardware Metrics
    # =============================================
    ws_hw = wb.create_sheet("Hardware Metrics")

    hw_headers = [
        "#", "Timestamp", "Server",
        "CPU (%)", "RAM Used (GB)", "RAM Total (GB)",
        "GPU (%)", "VRAM Used (GB)", "VRAM Total (GB)",
        "GPU Power (W)", "GPU Temp (°C)", "GPU Name",
        "Disk Read (MB/s)", "Disk Write (MB/s)",
    ]
    ws_hw.append(hw_headers)
    style_header_row(ws_hw, 1, len(hw_headers))

    for idx, h in enumerate(hardware, 1):
        ws_hw.append([
            idx,
            h.timestamp.strftime("%Y-%m-%d %H:%M:%S") if h.timestamp else "",
            h.server,
            round(h.cpu_pct, 1) if h.cpu_pct is not None else None,
            round(h.ram_used_gb, 2) if h.ram_used_gb is not None else None,
            round(h.ram_total_gb, 2) if h.ram_total_gb is not None else None,
            round(h.gpu_util_pct, 1) if h.gpu_util_pct is not None else None,
            round(h.vram_used_gb, 2) if h.vram_used_gb is not None else None,
            round(h.vram_total_gb, 2) if h.vram_total_gb is not None else None,
            round(h.gpu_power_watts, 1) if h.gpu_power_watts is not None else None,
            round(h.gpu_temperature_c, 1) if h.gpu_temperature_c is not None else None,
            h.gpu_name or "",
            round(h.disk_read_mbps, 2) if h.disk_read_mbps is not None else None,
            round(h.disk_write_mbps, 2) if h.disk_write_mbps is not None else None,
        ])

    for col_idx in range(1, len(hw_headers) + 1):
        ws_hw.column_dimensions[ws_hw.cell(1, col_idx).column_letter].width = max(
            len(str(hw_headers[col_idx - 1])) + 2, 12
        )

    # =============================================
    # Sheet 4: Server Comparisons (if ≥2 servers)
    # =============================================
    if comparisons:
        ws_comp = wb.create_sheet("Server Comparisons")

        comp_headers = [
            "#", "Tool", "Scenario", "Environment", "Concurrency",
            "S1 TTFT (ms)", "S1 TPS", "S1 RPS", "S1 P99 (ms)",
            "S2 TTFT (ms)", "S2 TPS", "S2 RPS", "S2 P99 (ms)",
            "Δ TTFT (%)", "Δ TPS (%)", "Δ RPS (%)", "Δ P99 (%)",
            "Winner",
        ]
        ws_comp.append(comp_headers)
        style_header_row(ws_comp, 1, len(comp_headers))

        for idx, c in enumerate(comparisons, 1):
            ws_comp.append([
                idx,
                c.tool or "all",
                c.scenario or "all",
                c.environment or "all",
                c.concurrency,
                round(c.s1_ttft_ms, 2) if c.s1_ttft_ms is not None else None,
                round(c.s1_tps, 2) if c.s1_tps is not None else None,
                round(c.s1_rps, 2) if c.s1_rps is not None else None,
                round(c.s1_p99_ms, 2) if c.s1_p99_ms is not None else None,
                round(c.s2_ttft_ms, 2) if c.s2_ttft_ms is not None else None,
                round(c.s2_tps, 2) if c.s2_tps is not None else None,
                round(c.s2_rps, 2) if c.s2_rps is not None else None,
                round(c.s2_p99_ms, 2) if c.s2_p99_ms is not None else None,
                round(c.delta_ttft_pct, 2) if c.delta_ttft_pct is not None else None,
                round(c.delta_tps_pct, 2) if c.delta_tps_pct is not None else None,
                round(c.delta_rps_pct, 2) if c.delta_rps_pct is not None else None,
                round(c.delta_p99_pct, 2) if c.delta_p99_pct is not None else None,
                c.overall_winner or "tie",
            ])

        for col_idx in range(1, len(comp_headers) + 1):
            ws_comp.column_dimensions[ws_comp.cell(1, col_idx).column_letter].width = max(
                len(str(comp_headers[col_idx - 1])) + 2, 12
            )

    # --- Write to buffer and return ---
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="aidaptive_report_{run_id}.xlsx"'
        },
    )

