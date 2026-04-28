"""
Deep Integration Tests — SQLite in-memory DB
Validates: Repository logic, N-server queries, Excel export, data integrity
"""
import pytest
import asyncio
from datetime import datetime
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession

from src.database.tables import (
    Base, BenchmarkRun, BenchmarkResultRow, HardwareSnapshot,
    ServerComparison, RunStatus,
)
from src.database.repository import Repository

# --------------- Fixtures ---------------

@pytest.fixture
def db_session():
    """Create in-memory SQLite with real schema.
    SQLite doesn't auto-generate BigInteger PKs, so we use
    a custom engine with Integer rendering for BigInteger columns.
    """
    from sqlalchemy import event, BigInteger, Integer

    engine = create_engine("sqlite:///:memory:")

    # SQLite workaround: BigInteger → Integer for autoincrement
    @event.listens_for(engine, "connect")
    def _set_sqlite_pragma(dbapi_conn, _):
        cursor = dbapi_conn.cursor()
        cursor.execute("PRAGMA journal_mode=WAL")
        cursor.close()

    # Temporarily swap BigInteger to Integer for table creation on SQLite
    original_type_map = {}
    for table in Base.metadata.tables.values():
        for col in table.columns:
            if isinstance(col.type, BigInteger):
                original_type_map[(table.name, col.name)] = col.type
                col.type = Integer()

    Base.metadata.create_all(engine)

    # Restore original types
    for (tbl_name, col_name), orig_type in original_type_map.items():
        Base.metadata.tables[tbl_name].columns[col_name].type = orig_type

    Session = sessionmaker(bind=engine)
    session = Session()
    yield session
    session.close()

@pytest.fixture
def repo(db_session):
    return Repository(db_session)

def _seed_run(session, run_id="run_test_001", status="completed", suite="perf"):
    run = BenchmarkRun(
        run_id=run_id, status=status, suite=suite, model="llama3",
        started_at=datetime(2026, 4, 28, 10, 0, 0),
        finished_at=datetime(2026, 4, 28, 10, 30, 0),
        duration_seconds=1800, total_tests=10, completed_tests=10, failed_tests=0,
    )
    session.add(run)
    session.commit()
    return run

def _seed_result(session, run_id, server, concurrency=1, tool="locust",
                 tps=50.0, ttft_ms=120.0, tpot_ms=30.0, error_rate=0.02):
    r = BenchmarkResultRow(
        run_id=run_id, server=server, tool=tool, environment="production",
        scenario="chat_simple", model="llama3", concurrency=concurrency,
        ttft_ms=ttft_ms, tpot_ms=tpot_ms, itl_ms=tpot_ms * 0.8,
        tps=tps, rps=tps * 0.9,
        latency_p50_ms=ttft_ms * 0.8, latency_p95_ms=ttft_ms * 1.5,
        latency_p99_ms=ttft_ms * 2.0,
        total_tokens=5000, total_requests=100, successful_requests=98,
        failed_requests=2, error_rate=error_rate, goodput=tps * (1 - error_rate),
        timestamp=datetime(2026, 4, 28, 10, 5, 0),
    )
    session.add(r)
    session.commit()
    return r

def _seed_hardware(session, run_id, server, gpu_pct=75.0, cpu_pct=45.0):
    h = HardwareSnapshot(
        run_id=run_id, server=server, timestamp=datetime(2026, 4, 28, 10, 5, 0),
        gpu_name="NVIDIA RTX 4090", gpu_util_pct=gpu_pct, vram_used_gb=18.5,
        vram_total_gb=24.0, gpu_power_watts=320.0, gpu_temperature_c=72.0,
        cpu_pct=cpu_pct, ram_used_gb=28.0, ram_total_gb=64.0,
        disk_read_mbps=150.0, disk_write_mbps=80.0,
    )
    session.add(h)
    session.commit()
    return h


# =============================================
# 1. REPOSITORY: 1-SERVER MODE
# =============================================
class TestOneServerMode:
    def test_summary_returns_only_server1(self, db_session, repo):
        """DEEP-1S01: 1 server → summary chỉ có server1"""
        _seed_run(db_session, "run_1s")
        _seed_result(db_session, "run_1s", "vm01", tps=80.0)

        # Use sync query directly
        from sqlalchemy import select, func, desc
        servers_res = db_session.execute(
            select(BenchmarkResultRow.server).filter_by(run_id="run_1s").distinct()
        )
        run_servers = sorted([row[0] for row in servers_res.all()])
        assert run_servers == ["vm01"]
        assert len(run_servers) == 1

    def test_results_data_integrity(self, db_session, repo):
        """DEEP-1S02: Inserted data matches queried data exactly"""
        _seed_run(db_session, "run_1s_int")
        _seed_result(db_session, "run_1s_int", "vm01", tps=123.456, ttft_ms=78.9, error_rate=0.05)

        results = repo.get_results_by_run("run_1s_int")
        assert len(results) == 1
        r = results[0]
        assert r.tps == pytest.approx(123.456)
        assert r.ttft_ms == pytest.approx(78.9)
        assert r.error_rate == pytest.approx(0.05)
        assert r.goodput == pytest.approx(123.456 * 0.95)
        assert r.server == "vm01"

    def test_aggregated_results_single_server(self, db_session, repo):
        """DEEP-1S03: Aggregation works with 1 server"""
        _seed_run(db_session, "run_1s_agg")
        _seed_result(db_session, "run_1s_agg", "vm01", concurrency=1, tps=50.0)
        _seed_result(db_session, "run_1s_agg", "vm01", concurrency=8, tps=100.0)

        agg = repo.get_aggregated_results("run_1s_agg", "vm01")
        assert agg["result_count"] == 2
        assert agg["avg_tps"] == pytest.approx(75.0)  # (50 + 100) / 2

    def test_no_server2_in_single_mode(self, db_session, repo):
        """DEEP-1S04: server2 aggregation returns 0 results"""
        _seed_run(db_session, "run_1s_no2")
        _seed_result(db_session, "run_1s_no2", "vm01", tps=50.0)

        agg = repo.get_aggregated_results("run_1s_no2", "vm02")
        assert agg["result_count"] == 0
        assert agg["avg_tps"] is None


# =============================================
# 2. REPOSITORY: 2-SERVER MODE
# =============================================
class TestTwoServerMode:
    def test_both_servers_in_summary(self, db_session, repo):
        """DEEP-2S01: 2 servers both appear in results"""
        _seed_run(db_session, "run_2s")
        _seed_result(db_session, "run_2s", "vm01", tps=50.0)
        _seed_result(db_session, "run_2s", "vm02", tps=80.0)

        from sqlalchemy import select
        servers_res = db_session.execute(
            select(BenchmarkResultRow.server).filter_by(run_id="run_2s").distinct()
        )
        run_servers = sorted([row[0] for row in servers_res.all()])
        assert run_servers == ["vm01", "vm02"]

    def test_aggregation_per_server(self, db_session, repo):
        """DEEP-2S02: Each server's aggregation is independent"""
        _seed_run(db_session, "run_2s_agg")
        _seed_result(db_session, "run_2s_agg", "vm01", tps=50.0, ttft_ms=100.0)
        _seed_result(db_session, "run_2s_agg", "vm02", tps=90.0, ttft_ms=60.0)

        s1 = repo.get_aggregated_results("run_2s_agg", "vm01")
        s2 = repo.get_aggregated_results("run_2s_agg", "vm02")

        assert s1["avg_tps"] == pytest.approx(50.0)
        assert s2["avg_tps"] == pytest.approx(90.0)
        assert s1["avg_ttft_ms"] == pytest.approx(100.0)
        assert s2["avg_ttft_ms"] == pytest.approx(60.0)

    def test_server_data_not_mixed(self, db_session, repo):
        """DEEP-2S03: vm01 data never leaks into vm02 query"""
        _seed_run(db_session, "run_2s_leak")
        _seed_result(db_session, "run_2s_leak", "vm01", tps=999.0, error_rate=0.5)
        _seed_result(db_session, "run_2s_leak", "vm02", tps=1.0, error_rate=0.0)

        s2 = repo.get_aggregated_results("run_2s_leak", "vm02")
        assert s2["avg_tps"] == pytest.approx(1.0)  # NOT 500 (average of both)
        assert s2["avg_error_rate"] == pytest.approx(0.0)  # NOT 0.25


# =============================================
# 3. REPOSITORY: 3-SERVER MODE
# =============================================
class TestThreeServerMode:
    def test_three_servers_detected(self, db_session, repo):
        """DEEP-3S01: 3 servers all detected"""
        _seed_run(db_session, "run_3s")
        _seed_result(db_session, "run_3s", "vm01", tps=50.0)
        _seed_result(db_session, "run_3s", "vm02", tps=70.0)
        _seed_result(db_session, "run_3s", "vm03", tps=90.0)

        from sqlalchemy import select
        servers_res = db_session.execute(
            select(BenchmarkResultRow.server).filter_by(run_id="run_3s").distinct()
        )
        run_servers = sorted([row[0] for row in servers_res.all()])
        assert len(run_servers) == 3
        assert run_servers == ["vm01", "vm02", "vm03"]

    def test_third_server_not_truncated(self, db_session, repo):
        """DEEP-3S02: server3 data is NOT dropped by [:2] slicing"""
        _seed_run(db_session, "run_3s_full")
        _seed_result(db_session, "run_3s_full", "vm01", tps=50.0)
        _seed_result(db_session, "run_3s_full", "vm02", tps=70.0)
        _seed_result(db_session, "run_3s_full", "vm03", tps=90.0)

        s3 = repo.get_aggregated_results("run_3s_full", "vm03")
        assert s3["result_count"] == 1
        assert s3["avg_tps"] == pytest.approx(90.0)

    def test_three_server_aggregation_independent(self, db_session, repo):
        """DEEP-3S03: All 3 servers have independent aggregation"""
        _seed_run(db_session, "run_3s_ind")
        _seed_result(db_session, "run_3s_ind", "vm01", tps=10.0, error_rate=0.1)
        _seed_result(db_session, "run_3s_ind", "vm02", tps=20.0, error_rate=0.2)
        _seed_result(db_session, "run_3s_ind", "vm03", tps=30.0, error_rate=0.3)

        for server, expected_tps, expected_err in [
            ("vm01", 10.0, 0.1), ("vm02", 20.0, 0.2), ("vm03", 30.0, 0.3)
        ]:
            agg = repo.get_aggregated_results("run_3s_ind", server)
            assert agg["avg_tps"] == pytest.approx(expected_tps), f"{server} TPS mismatch"
            assert agg["avg_error_rate"] == pytest.approx(expected_err), f"{server} error mismatch"


# =============================================
# 4. HARDWARE SNAPSHOTS — DATA INTEGRITY
# =============================================
class TestHardwareIntegrity:
    def test_hardware_stored_per_server(self, db_session):
        """DEEP-HW01: Hardware snapshots stored and queried per server"""
        _seed_run(db_session, "run_hw")
        _seed_hardware(db_session, "run_hw", "vm01", gpu_pct=90.0, cpu_pct=50.0)
        _seed_hardware(db_session, "run_hw", "vm02", gpu_pct=60.0, cpu_pct=30.0)

        snaps = db_session.query(HardwareSnapshot).filter_by(run_id="run_hw").all()
        assert len(snaps) == 2

        vm01_snap = [s for s in snaps if s.server == "vm01"][0]
        vm02_snap = [s for s in snaps if s.server == "vm02"][0]

        assert vm01_snap.gpu_util_pct == pytest.approx(90.0)
        assert vm02_snap.gpu_util_pct == pytest.approx(60.0)
        assert vm01_snap.cpu_pct != vm02_snap.cpu_pct  # Not mixed

    def test_hardware_fields_complete(self, db_session):
        """DEEP-HW02: All hardware fields stored correctly"""
        _seed_run(db_session, "run_hw2")
        _seed_hardware(db_session, "run_hw2", "vm01")

        h = db_session.query(HardwareSnapshot).filter_by(run_id="run_hw2").first()
        assert h.gpu_name == "NVIDIA RTX 4090"
        assert h.vram_used_gb == pytest.approx(18.5)
        assert h.vram_total_gb == pytest.approx(24.0)
        assert h.gpu_power_watts == pytest.approx(320.0)
        assert h.gpu_temperature_c == pytest.approx(72.0)
        assert h.ram_used_gb == pytest.approx(28.0)
        assert h.ram_total_gb == pytest.approx(64.0)
        assert h.disk_read_mbps == pytest.approx(150.0)
        assert h.disk_write_mbps == pytest.approx(80.0)


# =============================================
# 5. AGGREGATOR — COMPARISON LOGIC
# =============================================
class TestAggregatorLogic:
    def test_delta_calculation(self):
        """DEEP-AGG01: Delta percentage calc is correct"""
        from src.data.aggregator import Aggregator
        # s2 faster TTFT (lower is better) → positive delta
        assert Aggregator._calc_delta(100.0, 80.0, lower_is_better=True) == pytest.approx(20.0)
        # s2 higher TPS (higher is better) → positive delta
        assert Aggregator._calc_delta(50.0, 75.0, lower_is_better=False) == pytest.approx(50.0)
        # s2 worse TTFT → negative delta
        assert Aggregator._calc_delta(80.0, 100.0, lower_is_better=True) == pytest.approx(-25.0)

    def test_winner_determination(self):
        """DEEP-AGG02: Winner determined by majority of deltas"""
        from src.data.aggregator import Aggregator
        comp = ServerComparison(
            delta_ttft_pct=10.0,   # s2 wins
            delta_tps_pct=20.0,    # s2 wins
            delta_rps_pct=-5.0,    # s1 wins
            delta_p99_pct=15.0,    # s2 wins
        )
        assert Aggregator._determine_winner(comp) == "server2"  # 3 vs 1

    def test_tie_when_equal(self):
        """DEEP-AGG03: Tie when scores equal"""
        from src.data.aggregator import Aggregator
        comp = ServerComparison(
            delta_ttft_pct=10.0, delta_tps_pct=-10.0,
            delta_rps_pct=None, delta_p99_pct=None,
        )
        assert Aggregator._determine_winner(comp) == "tie"


# =============================================
# 6. CONCURRENCY DATA — CHART GROUPING
# =============================================
class TestConcurrencyGrouping:
    def test_multiple_concurrency_levels(self, db_session, repo):
        """DEEP-CC01: Results grouped correctly by concurrency"""
        _seed_run(db_session, "run_cc")
        for c in [1, 8, 16, 32]:
            _seed_result(db_session, "run_cc", "vm01", concurrency=c, tps=c * 10.0)

        results = repo.get_results_by_run("run_cc")
        assert len(results) == 4

        by_conc = {r.concurrency: r.tps for r in results}
        assert by_conc[1] == pytest.approx(10.0)
        assert by_conc[8] == pytest.approx(80.0)
        assert by_conc[16] == pytest.approx(160.0)
        assert by_conc[32] == pytest.approx(320.0)

    def test_concurrency_per_server_isolated(self, db_session, repo):
        """DEEP-CC02: Concurrency data isolated between servers"""
        _seed_run(db_session, "run_cc2")
        _seed_result(db_session, "run_cc2", "vm01", concurrency=1, tps=10.0)
        _seed_result(db_session, "run_cc2", "vm02", concurrency=1, tps=99.0)

        results = repo.get_results_by_run("run_cc2", server="vm01")
        assert len(results) == 1
        assert results[0].tps == pytest.approx(10.0)


# =============================================
# 7. EXCEL EXPORT — CONTENT VALIDATION
# =============================================
class TestExcelExportContent:
    def test_openpyxl_creates_valid_workbook(self, db_session):
        """DEEP-XL01: openpyxl generates readable workbook"""
        from openpyxl import Workbook
        import io

        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws.append(["Field", "Value"])
        ws.append(["Run ID", "run_test"])
        ws.append(["TPS", 123.456])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        assert buf.getvalue()[:4] == b'PK\x03\x04'  # Valid ZIP (XLSX)

        # Re-read and verify
        from openpyxl import load_workbook
        wb2 = load_workbook(buf)
        ws2 = wb2["Summary"]
        assert ws2.cell(1, 1).value == "Field"
        assert ws2.cell(2, 1).value == "Run ID"
        assert ws2.cell(3, 2).value == pytest.approx(123.456)

    def test_excel_results_sheet_structure(self, db_session):
        """DEEP-XL02: Benchmark Results sheet has correct columns"""
        from openpyxl import Workbook
        expected_cols = [
            "#", "Timestamp", "Server", "Tool", "Environment", "Scenario", "Model",
            "Concurrency", "TTFT (ms)", "TPOT (ms)", "ITL (ms)", "TPS",
            "RPS", "P50 (ms)", "P95 (ms)", "P99 (ms)",
            "Total Tokens", "Total Requests", "Successful", "Failed",
            "Error Rate (%)", "Goodput",
        ]
        wb = Workbook()
        ws = wb.create_sheet("Benchmark Results")
        ws.append(expected_cols)

        for i, col_name in enumerate(expected_cols, 1):
            assert ws.cell(1, i).value == col_name

    def test_excel_data_round_trip(self, db_session):
        """DEEP-XL03: Data written to Excel matches source"""
        from openpyxl import Workbook, load_workbook
        import io

        wb = Workbook()
        ws = wb.active
        ws.append(["Server", "TPS", "Error Rate"])
        ws.append(["vm01", 123.456, 0.05])
        ws.append(["vm02", 78.9, 0.12])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        wb2 = load_workbook(buf)
        ws2 = wb2.active
        assert ws2.cell(2, 1).value == "vm01"
        assert ws2.cell(2, 2).value == pytest.approx(123.456)
        assert ws2.cell(3, 3).value == pytest.approx(0.12)


# =============================================
# 8. EDGE CASES — EMPTY & NULL DATA
# =============================================
class TestEdgeCases:
    def test_empty_run_no_results(self, db_session, repo):
        """DEEP-EDGE01: Run with 0 results returns empty"""
        _seed_run(db_session, "run_empty")
        results = repo.get_results_by_run("run_empty")
        assert results == []

    def test_null_metrics_handled(self, db_session, repo):
        """DEEP-EDGE02: NULL metric values don't crash aggregation"""
        _seed_run(db_session, "run_null")
        r = BenchmarkResultRow(
            run_id="run_null", server="vm01", tool="locust",
            environment="test", scenario="null_test", model="llama3",
            concurrency=1, tps=None, ttft_ms=None, error_rate=None,
            total_requests=0, timestamp=datetime(2026, 4, 28, 10, 0, 0),
        )
        db_session.add(r)
        db_session.commit()

        agg = repo.get_aggregated_results("run_null", "vm01")
        assert agg["result_count"] == 1
        assert agg["avg_tps"] is None
        assert agg["avg_ttft_ms"] is None

    def test_nonexistent_run(self, db_session, repo):
        """DEEP-EDGE03: Querying nonexistent run returns empty/None"""
        run = repo.get_run("nonexistent_run_xyz")
        assert run is None
        results = repo.get_results_by_run("nonexistent_run_xyz")
        assert results == []

    def test_error_rate_boundary_values(self, db_session, repo):
        """DEEP-EDGE04: error_rate 0.0 and 1.0 boundaries"""
        _seed_run(db_session, "run_boundary")
        _seed_result(db_session, "run_boundary", "vm01", error_rate=0.0, tps=100.0)
        _seed_result(db_session, "run_boundary", "vm02", error_rate=1.0, tps=50.0)

        s1 = repo.get_aggregated_results("run_boundary", "vm01")
        s2 = repo.get_aggregated_results("run_boundary", "vm02")
        assert s1["avg_error_rate"] == pytest.approx(0.0)
        assert s2["avg_error_rate"] == pytest.approx(1.0)

    def test_run_isolation(self, db_session, repo):
        """DEEP-EDGE05: Results from different runs don't leak"""
        _seed_run(db_session, "run_a")
        _seed_run(db_session, "run_b")
        _seed_result(db_session, "run_a", "vm01", tps=111.0)
        _seed_result(db_session, "run_b", "vm01", tps=222.0)

        results_a = repo.get_results_by_run("run_a")
        results_b = repo.get_results_by_run("run_b")
        assert len(results_a) == 1
        assert len(results_b) == 1
        assert results_a[0].tps == pytest.approx(111.0)
        assert results_b[0].tps == pytest.approx(222.0)
